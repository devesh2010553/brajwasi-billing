from flask import Flask, render_template, request, redirect, session, send_from_directory
import json, os, math, calendar
from datetime import datetime, timedelta, time
from google.oauth2 import service_account
from googleapiclient.discovery import build

app = Flask(__name__)
app.secret_key = "supersecretkey"
ADMIN_CODE = os.getenv("ADMIN_CODE", "admin1234")   # set ADMIN_CODE env var to change

# ---------- SESSION LIFETIME (10 YEARS) ----------
app.permanent_session_lifetime = timedelta(days=3650)

# ---------- Load drivers ----------
with open("driver.json", "r") as f:
    DRIVERS = json.load(f)

# ---------- Google Auth ----------
sa_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
if not sa_json:
    raise RuntimeError("GOOGLE_SERVICE_ACCOUNT_JSON not set")

creds = service_account.Credentials.from_service_account_info(
    json.loads(sa_json),
    scopes=[
        "https://www.googleapis.com/auth/spreadsheets"
    ]
)

sheets = build("sheets", "v4", credentials=creds)

# ---------- Helpers ----------
def today_date():
    return datetime.now().date()

def parse_time(t):
    return datetime.strptime(t, "%H:%M").time()

def hours_between(start, end):
    d1 = datetime.combine(today_date(), start)
    d2 = datetime.combine(today_date(), end)
    if d2 < d1:
        d2 += timedelta(days=1)
    return (d2 - d1).total_seconds() / 3600

def calculate_ot(start, end):
    hrs = hours_between(start, end)
    extra = hrs - 12
    if extra <= 0:
        return 0
    full_hours = int(extra)          # whole hours of extra time
    fraction = extra - full_hours    # remaining minutes as fraction of hour
    if fraction > 0.5:
        return full_hours + 1        # e.g. 1h 35min → 2
    elif full_hours == 0:
        return 0                     # e.g. 0h 25min → 0 (under half hour)
    else:
        return full_hours            # e.g. 1h 25min → 1

def get_remarks(start, end, date):
    night_start = start < time(5, 0)       # started before 5 AM
    night_end = end >= time(22, 0)          # closed at or after 10 PM
    sunday = date.weekday() == 6

    parts = []

    if night_start and night_end:
        parts.append("Night Night")
    elif night_start or night_end:
        parts.append("Night")

    if sunday:
        parts.append("Sunday")

    return "/".join(parts)

# ---------- PWA Routes ----------
@app.route('/manifest.json')
def manifest():
    return send_from_directory(os.getcwd(), 'manifest.json')

@app.route('/service-worker.js')
def sw():
    return send_from_directory(os.getcwd(), 'service-worker.js')

# ---------- Routes ----------
@app.route("/", methods=["GET", "POST"])
def login():

    # Skip login if already logged in
    if "car" in session:
        return redirect("/entry")

    msg = ""

    if request.method == "POST":
        code = request.form["code"]

        for car, info in DRIVERS.items():
            if info["code"] == code:

                session.permanent = True   # keeps login for 10 years
                session["car"] = car

                return redirect("/entry")

        msg = "Invalid code"

    return render_template("login.html", msg=msg)

@app.route("/entry", methods=["GET", "POST"])
def entry():
    if "car" not in session:
        return redirect("/")

    car = session["car"]
    info = DRIVERS[car]
    msg = ""
    cls = "success"

    if request.method == "POST":
        try:
            opening = int(request.form["opening"])
            closing = int(request.form["closing"])
            start = parse_time(request.form["start"])
            end = parse_time(request.form["end"])

            entry_date_str = request.form.get("entry_date", "")
            entry_date = datetime.strptime(entry_date_str, "%Y-%m-%d").date() if entry_date_str else today_date()

            remarks = get_remarks(start, end, entry_date)
            ot = calculate_ot(start, end)

            row = entry_date.day + 7
            rng = f"{info['sheet']}!C{row}:I{row}"

            values = [[
                opening,
                closing,
                closing - opening,
                start.strftime("%I:%M %p"),
                end.strftime("%I:%M %p"),
                ot,
                remarks
            ]]

            sheets.spreadsheets().values().update(
                spreadsheetId=info["file_id"],
                range=rng,
                valueInputOption="USER_ENTERED",
                body={"values": values}
            ).execute()

            msg = "Saved successfully"

        except Exception as e:
            msg = str(e)
            cls = "error"

    return render_template("entry.html", car=car, msg=msg, cls=cls, today=today_date().isoformat())

@app.route("/check-entry", methods=["POST"])
def check_entry():
    if "car" not in session:
        return {"filled": False}

    car = session["car"]
    info = DRIVERS[car]

    try:
        entry_date_str = request.json.get("entry_date", "")
        entry_date = datetime.strptime(entry_date_str, "%Y-%m-%d").date()
        row = entry_date.day + 7
        rng = f"{info['sheet']}!C{row}"

        result = sheets.spreadsheets().values().get(
            spreadsheetId=info["file_id"],
            range=rng
        ).execute()

        values = result.get("values", [])
        filled = bool(values and values[0] and str(values[0][0]).strip() != "")
        return {"filled": filled}
    except Exception as e:
        return {"filled": False, "error": str(e)}

@app.route("/get-last-closing", methods=["POST"])
def get_last_closing():
    if "car" not in session:
        return {"closing": None}

    car = session["car"]
    info = DRIVERS[car]

    try:
        entry_date_str = request.json.get("entry_date", "")
        entry_date = datetime.strptime(entry_date_str, "%Y-%m-%d").date()

        # Go back day by day (up to 7 days) to find last filled closing KM
        for i in range(1, 8):
            prev_date = entry_date - timedelta(days=i)
            prev_row = prev_date.day + 7
            # Column D is closing KM (C=opening, D=closing)
            rng = f"{info['sheet']}!D{prev_row}"
            result = sheets.spreadsheets().values().get(
                spreadsheetId=info["file_id"],
                range=rng
            ).execute()
            values = result.get("values", [])
            if values and values[0] and str(values[0][0]).strip() != "":
                return {"closing": values[0][0]}

        return {"closing": None}
    except Exception as e:
        return {"closing": None, "error": str(e)}

def ordinal_suffix(n):
    if 11 <= n <= 13:
        return "th"
    return {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")

def ordinal(n):
    return f"{n}{ordinal_suffix(n)}"

# ── DRIVER MASTER CONFIG ─────────────────────────────────────────────────────
# Each entry maps car key → static info used when building Excel files
DRIVER_META = {
    "UP80JT4509": {"user": "Mr. Ghanshyam (S&T)", "type": "Ertiga", "location": "Tudla",    "amount": 55000},
    "UP80JT7912": {"user": "Mr. Raju Sen(S&T)",   "type": "Ertiga", "location": "Aligarh",  "amount": 55000},
    "UP79AT9051": {"user": "Mr. Abhishek (S&T)",  "type": "Ertiga", "location": "Divyapur", "amount": 55000},
    "UP80JT5885": {"user": "Driver 5885",          "type": "Ertiga", "location": "Location", "amount": 55000},
    "UP80JT6593": {"user": "Driver 6593",          "type": "Ertiga", "location": "Location", "amount": 51750},
}
FILE_GROUPS = {
    "S_T_BT": ["UP80JT4509", "UP80JT7912", "UP79AT9051"],
    "BT":     ["UP80JT5885", "UP80JT6593"],
}

# ── EXCEL BUILDER ─────────────────────────────────────────────────────────────
def build_excel_files(month, year, sheet_data_map):
    """
    Build both Excel files for the given month/year.
    sheet_data_map: {reg_no: [[row_values], ...]} — actual data from Google Sheets
    Returns dict of {group_name: filepath}
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import date as dt_date

    days   = calendar.monthrange(year, month)[1]
    mname  = datetime(year, month, 1).strftime("%B")
    title  = (f" Vehicle Bill for the period from {ordinal(1)} {mname} {year}"
              f" to {ordinal(days)} {mname} {year}")

    def yf():   return PatternFill("solid", fgColor="FFFFFF00")
    def nof():  return PatternFill(fill_type=None)
    def fn(bold=False, size=11): return Font(bold=bold, size=size)
    def al(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def sd(s): return Side(style=s) if s else Side(style=None)
    def brd(l=None,r=None,t=None,b=None):
        return Border(left=sd(l),right=sd(r),top=sd(t),bottom=sd(b))
    def athin(): return brd("thin","thin","thin","thin")

    def build_sheet(wb, reg):
        meta = DRIVER_META.get(reg, {})
        ws   = wb.create_sheet(title=reg)
        widths = {'A':8.71,'B':10.14,'C':10.29,'D':9.86,'E':8.14,
                  'F':9.0,'G':8.86,'H':12.71,'I':13.86}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.row_dimensions[1].height = 21.75
        ws.row_dimensions[6].height = 28.5

        def mc(r1,c1,r2,c2): ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
        def sc(coord, val, bold=False, size=11, h="center", wrap=False, fill=None, brdr=None):
            c = ws[coord]; c.value = val
            c.font = fn(bold,size); c.alignment = al(h)
            if fill: c.fill = fill
            if brdr: c.border = brdr

        mc(1,1,1,9); sc("A1","BRAJWASI TRAVELS",bold=True,size=16)
        mc(2,1,2,9); sc("A2","Plot No. 42 Deep Nagar Phase-2, Dehtora Road, Bodla, Agar, 282007",bold=True)
        mc(3,1,3,9); sc("A3",title,bold=True)

        ws["A4"].value=f"Vehicle Reg. No.  :  {reg}"; ws["A4"].font=fn(True); ws["A4"].alignment=al(h="left")
        mc(4,5,4,9); sc("E4",f"      User Name / Dept. : {meta.get('user','')}",bold=True)
        mc(5,1,5,3); ws["A5"].value=f"Vehicle type : {meta.get('type','Ertiga')}"; ws["A5"].font=fn(True); ws["A5"].alignment=al(h="left")
        mc(5,6,5,9); sc("F5",f"Slice / Location :  {meta.get('location','')}",bold=True)

        mc(6,1,7,1); sc("A6","SL.NO.",bold=True,brdr=brd("medium","thin","thin","thin"))
        mc(6,2,7,2); sc("B6","DATE",bold=True,brdr=athin())
        mc(6,3,6,4); sc("C6","READING (KM)",bold=True,brdr=athin())
        sc("E6","TOTAL RUNS",bold=True,brdr=brd("thin","thin","thin",None))
        mc(6,6,6,7); sc("F6","TIMING",bold=True,brdr=athin())
        sc("H6","OT. HRS",bold=True,brdr=brd("thin","thin","thin",None))
        mc(6,9,7,9); sc("I6","REMARKS",bold=True,brdr=brd("thin","medium","thin","thin"))

        for coord,val,sz in [("C7","OPENING",11),("D7","CLOSING",11),("E7","(KM)",11),
                               ("F7","OPENING",10),("G7","CLOSING",10),("H7","(After 12 hrs)",10)]:
            ws[coord].value=val; ws[coord].font=fn(True,sz); ws[coord].alignment=al()
            ws[coord].border=athin()
        ws["A7"].border=brd("medium","thin",None,"thin")
        ws["B7"].border=athin()
        ws["I7"].border=brd("thin","medium",None,"thin")

        # Data rows
        row_data = sheet_data_map.get(reg, [])
        for day in range(1, days+1):
            r = day + 7
            d = dt_date(year, month, day)
            ws.cell(r,1).value=day; ws.cell(r,1).alignment=al(); ws.cell(r,1).border=brd("medium","thin","thin","thin")
            ws.cell(r,2).value=d; ws.cell(r,2).number_format="d-mmm-yy"
            ws.cell(r,2).alignment=al(); ws.cell(r,2).fill=yf(); ws.cell(r,2).border=athin()

            # Fill in actual data if available (index: day-1, cols C-I = indices 0-6)
            if day-1 < len(row_data) and row_data[day-1]:
                vals = row_data[day-1]
                for ci, vi in enumerate(vals):
                    ws.cell(r, ci+3).value = vi if vi != '' else None

            for col in range(3,9):
                c = ws.cell(r,col)
                c.alignment=al()
                if col==6: c.number_format='[$-409]h:mm\\ AM/PM;@'
                elif col==7: c.number_format='[$-409]h:mm\\ AM/PM;@'
                elif col==8: c.number_format='0.00'
                c.border=brd("medium" if col==1 else "thin",
                             "medium" if col==8 else "thin","thin","thin")
            ws.cell(r,9).alignment=al(); ws.cell(r,9).border=brd("thin","medium","thin","thin")

        # Total row
        tr = days+8
        mc(tr,1,tr,2); ws.cell(tr,1).border=brd("medium","thin","thin","thin")
        mc(tr,3,tr,4)
        ws.cell(tr,3).value="TOTAL:"; ws.cell(tr,3).alignment=al(); ws.cell(tr,3).border=athin()
        ws.cell(tr,5).value=f"=SUM(E8:E{tr-1})"; ws.cell(tr,5).fill=yf(); ws.cell(tr,5).alignment=al(); ws.cell(tr,5).border=athin()
        for col in [4,6,7]: ws.cell(tr,col).border=athin()
        ws.cell(tr,8).value=f"=SUM(H8:H{tr-1})"; ws.cell(tr,8).fill=yf(); ws.cell(tr,8).alignment=al(); ws.cell(tr,8).border=athin()
        ws.cell(tr,9).border=brd("thin","medium","thin","thin")

        # Payment section
        pr = tr+1
        mc(pr,1,pr,9); ws.cell(pr,1).value="Payment Details      "; ws.cell(pr,1).alignment=al()

        amt = meta.get("amount", 55000)
        pay_rows = [
            ("Amount for 30 Days per 3000 K.M.", 3000, 30, amt),
            ("Extra Kms",    0,  12, f"=E{pr+2}*F{pr+2}"),
            ("OT hrs",       f"=H{tr}", 50, f"=E{pr+3}*F{pr+3}"),
            ("Sunday",       0, 500, f"=E{pr+4}*F{pr+4}"),
            ("Holiday",      0, 500, f"=E{pr+5}*F{pr+5}"),
            ("Night charge", 0, 400, f"=E{pr+6}*F{pr+6}"),
        ]
        for i,(lbl,ev,fv,gv) in enumerate(pay_rows):
            row = pr+1+i
            mc(row,1,row,4)
            ws.cell(row,1).value=lbl; ws.cell(row,1).font=fn(True); ws.cell(row,1).alignment=al()
            ws.cell(row,5).value=ev; ws.cell(row,5).alignment=al()
            ws.cell(row,6).value=fv; ws.cell(row,6).alignment=al()
            cg=ws.cell(row,7); cg.value=gv; cg.alignment=al()
            if isinstance(gv,str) and gv.startswith("="): cg.fill=yf()

        toll_r = pr+7
        mc(toll_r,1,toll_r,4); ws.cell(toll_r,1).value="Toll Tax"; ws.cell(toll_r,1).font=fn(True)
        grand_r = pr+8
        mc(grand_r,1,grand_r,4)
        ws.cell(grand_r,1).value="Grand Total"; ws.cell(grand_r,1).font=fn(True); ws.cell(grand_r,1).alignment=al()
        cgt=ws.cell(grand_r,7); cgt.value=f"=SUM(G{pr+1}:G{grand_r-1})"; cgt.fill=yf(); cgt.alignment=al()

    results = {}
    os.makedirs("temp_exports", exist_ok=True)
    for group, regs in FILE_GROUPS.items():
        wb = Workbook()
        wb.remove(wb.active)
        for reg in regs:
            with open("driver.json") as f:
                drivers = json.load(f)
            # Find the reg in drivers
            reg_in_drivers = any(info.get("sheet") == reg for info in drivers.values())
            if reg_in_drivers:
                build_sheet(wb, reg)
        if len(wb.sheetnames) > 0:
            fname = f"temp_exports/{group}_{mname}_{year}.xlsx"
            wb.save(fname)
            results[group] = fname
    return results

def fetch_sheet_data(file_id, sheet_name, days):
    """Fetch all data rows C:I from Google Sheets for a given sheet."""
    rng = f"{sheet_name}!C8:I{7+days}"
    result = sheets.spreadsheets().values().get(
        spreadsheetId=file_id, range=rng
    ).execute()
    rows = result.get("values", [])
    # Pad to full days length
    padded = []
    for day in range(days):
        if day < len(rows) and rows[day]:
            padded.append(rows[day])
        else:
            padded.append([])
    return padded

def reset_sheet_for_new_month(file_id, sheet_name, month, year):
    """Reset the Google Sheet tab for the new month."""
    days = calendar.monthrange(year, month)[1]
    mname = datetime(year, month, 1).strftime("%B")
    new_title = (f" Vehicle Bill for the period from {ordinal(1)} {mname} {year}"
                 f" to {ordinal(days)} {mname} {year}")

    sheets.spreadsheets().values().update(
        spreadsheetId=file_id, range=f"{sheet_name}!A3",
        valueInputOption="USER_ENTERED", body={"values": [[new_title]]}
    ).execute()

    date_values = [
        [day, datetime(year, month, day).strftime("%d-%b-%y")]
        for day in range(1, days+1)
    ]
    sheets.spreadsheets().values().update(
        spreadsheetId=file_id, range=f"{sheet_name}!A8:B{7+days}",
        valueInputOption="USER_ENTERED", body={"values": date_values}
    ).execute()

    if days < 31:
        sheets.spreadsheets().values().clear(
            spreadsheetId=file_id,
            range=f"{sheet_name}!A{8+days}:I{7+31}"
        ).execute()

    sheets.spreadsheets().values().clear(
        spreadsheetId=file_id,
        range=f"{sheet_name}!C8:I{7+days}"
    ).execute()

@app.route("/admin", methods=["GET", "POST"])
def admin():
    msg = ""
    cls = "error"
    download_links = []

    if request.method == "POST":
        code = request.form.get("code", "")
        if code != ADMIN_CODE:
            msg = "Invalid admin code"
        else:
            try:
                month      = int(request.form["month"])
                year       = int(request.form["year"])
                month_name = datetime(year, month, 1).strftime("%B")

                with open("driver.json") as f:
                    drivers = json.load(f)

                # Step 1: Fetch current (old) month data from all sheets
                prev_month_date = datetime(year, month, 1) - timedelta(days=1)
                prev_month = prev_month_date.month
                prev_year  = prev_month_date.year
                prev_days  = calendar.monthrange(prev_year, prev_month)[1]

                sheet_data_map = {}
                for car, info in drivers.items():
                    reg = info["sheet"]
                    sheet_data_map[reg] = fetch_sheet_data(info["file_id"], reg, prev_days)

                # Step 2: Build old-month Excel files with fetched data
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from datetime import date as dt_date

                prev_mname = prev_month_date.strftime("%B")
                prev_days_in = calendar.monthrange(prev_year, prev_month)[1]
                prev_title = (f" Vehicle Bill for the period from {ordinal(1)} {prev_mname} {prev_year}"
                              f" to {ordinal(prev_days_in)} {prev_mname} {prev_year}")

                # Temporarily swap DRIVER_META title reference for old month
                # We build files using the helper inline here
                os.makedirs("temp_exports", exist_ok=True)
                export_files = {}

                for group, regs in FILE_GROUPS.items():
                    wb = Workbook()
                    if wb.active: wb.remove(wb.active)

                    for reg in regs:
                        if not any(info.get("sheet") == reg for info in drivers.values()):
                            continue
                        meta = DRIVER_META.get(reg, {})
                        ws   = wb.create_sheet(title=reg)

                        def fn(bold=False,sz=11): return Font(bold=bold,size=sz)
                        def al(h="center",v="center",wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
                        def sd(s): return Side(style=s) if s else Side(style=None)
                        def brd(l=None,r=None,t=None,b=None): return Border(left=sd(l),right=sd(r),top=sd(t),bottom=sd(b))
                        def athin(): return brd("thin","thin","thin","thin")
                        def yf(): return PatternFill("solid",fgColor="FFFFFF00")

                        widths={'A':8.71,'B':10.14,'C':10.29,'D':9.86,'E':8.14,'F':9.0,'G':8.86,'H':12.71,'I':13.86}
                        for col,w in widths.items(): ws.column_dimensions[col].width=w
                        ws.row_dimensions[1].height=21.75; ws.row_dimensions[6].height=28.5

                        def mc(r1,c1,r2,c2): ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
                        def sc(coord,val,bold=False,sz=11,h="center",brdr=None,fill=None):
                            c=ws[coord]; c.value=val; c.font=fn(bold,sz); c.alignment=al(h)
                            if brdr: c.border=brdr
                            if fill: c.fill=fill

                        mc(1,1,1,9); sc("A1","BRAJWASI TRAVELS",bold=True,sz=16)
                        mc(2,1,2,9); sc("A2","Plot No. 42 Deep Nagar Phase-2, Dehtora Road, Bodla, Agar, 282007",bold=True)
                        mc(3,1,3,9); sc("A3",prev_title,bold=True)
                        ws["A4"].value=f"Vehicle Reg. No.  :  {reg}"; ws["A4"].font=fn(True); ws["A4"].alignment=al(h="left")
                        mc(4,5,4,9); sc("E4",f"      User Name / Dept. : {meta.get('user','')}",bold=True)
                        mc(5,1,5,3); ws["A5"].value=f"Vehicle type : {meta.get('type','Ertiga')}"; ws["A5"].font=fn(True); ws["A5"].alignment=al(h="left")
                        mc(5,6,5,9); sc("F5",f"Slice / Location :  {meta.get('location','')}",bold=True)

                        mc(6,1,7,1); sc("A6","SL.NO.",bold=True,brdr=brd("medium","thin","thin","thin"))
                        mc(6,2,7,2); sc("B6","DATE",bold=True,brdr=athin())
                        mc(6,3,6,4); sc("C6","READING (KM)",bold=True,brdr=athin())
                        sc("E6","TOTAL RUNS",bold=True,brdr=brd("thin","thin","thin",None))
                        mc(6,6,6,7); sc("F6","TIMING",bold=True,brdr=athin())
                        sc("H6","OT. HRS",bold=True,brdr=brd("thin","thin","thin",None))
                        mc(6,9,7,9); sc("I6","REMARKS",bold=True,brdr=brd("thin","medium","thin","thin"))
                        for coord,val,sz in [("C7","OPENING",11),("D7","CLOSING",11),("E7","(KM)",11),("F7","OPENING",10),("G7","CLOSING",10),("H7","(After 12 hrs)",10)]:
                            ws[coord].value=val; ws[coord].font=fn(True,sz); ws[coord].alignment=al(); ws[coord].border=athin()
                        ws["A7"].border=brd("medium","thin",None,"thin"); ws["B7"].border=athin(); ws["I7"].border=brd("thin","medium",None,"thin")

                        row_data = sheet_data_map.get(reg, [])
                        for day in range(1, prev_days_in+1):
                            r=day+7; d=dt_date(prev_year,prev_month,day)
                            ws.cell(r,1).value=day; ws.cell(r,1).alignment=al(); ws.cell(r,1).border=brd("medium","thin","thin","thin")
                            ws.cell(r,2).value=d; ws.cell(r,2).number_format="d-mmm-yy"
                            ws.cell(r,2).alignment=al(); ws.cell(r,2).fill=yf(); ws.cell(r,2).border=athin()
                            if day-1 < len(row_data) and row_data[day-1]:
                                vals=row_data[day-1]
                                for ci,vi in enumerate(vals): ws.cell(r,ci+3).value=vi if vi!='' else None
                            for col in range(3,9):
                                c=ws.cell(r,col); c.alignment=al()
                                if col==6: c.number_format='[$-409]h:mm\\ AM/PM;@'
                                elif col==7: c.number_format='[$-409]h:mm\\ AM/PM;@'
                                elif col==8: c.number_format='0.00'
                                c.border=brd("medium" if col==1 else "thin","medium" if col==8 else "thin","thin","thin")
                            ws.cell(r,9).alignment=al(); ws.cell(r,9).border=brd("thin","medium","thin","thin")

                        tr=prev_days_in+8
                        mc(tr,1,tr,2); ws.cell(tr,1).border=brd("medium","thin","thin","thin")
                        mc(tr,3,tr,4); ws.cell(tr,3).value="TOTAL:"; ws.cell(tr,3).alignment=al(); ws.cell(tr,3).border=athin()
                        ws.cell(tr,5).value=f"=SUM(E8:E{tr-1})"; ws.cell(tr,5).fill=yf(); ws.cell(tr,5).alignment=al(); ws.cell(tr,5).border=athin()
                        for col in [4,6,7]: ws.cell(tr,col).border=athin()
                        ws.cell(tr,8).value=f"=SUM(H8:H{tr-1})"; ws.cell(tr,8).fill=yf(); ws.cell(tr,8).alignment=al(); ws.cell(tr,8).border=athin()
                        ws.cell(tr,9).border=brd("thin","medium","thin","thin")

                        pr=tr+1; mc(pr,1,pr,9); ws.cell(pr,1).value="Payment Details      "; ws.cell(pr,1).alignment=al()
                        pay_rows=[
                            ("Amount for 30 Days per 3000 K.M.",3000,30,meta.get("amount",55000)),
                            ("Extra Kms",0,12,f"=E{pr+2}*F{pr+2}"),
                            ("OT hrs",f"=H{tr}",50,f"=E{pr+3}*F{pr+3}"),
                            ("Sunday",0,500,f"=E{pr+4}*F{pr+4}"),
                            ("Holiday",0,500,f"=E{pr+5}*F{pr+5}"),
                            ("Night charge",0,400,f"=E{pr+6}*F{pr+6}"),
                        ]
                        for i,(lbl,ev,fv,gv) in enumerate(pay_rows):
                            row=pr+1+i; mc(row,1,row,4)
                            ws.cell(row,1).value=lbl; ws.cell(row,1).font=fn(True); ws.cell(row,1).alignment=al()
                            ws.cell(row,5).value=ev; ws.cell(row,5).alignment=al()
                            ws.cell(row,6).value=fv; ws.cell(row,6).alignment=al()
                            cg=ws.cell(row,7); cg.value=gv; cg.alignment=al()
                            if isinstance(gv,str) and gv.startswith("="): cg.fill=yf()
                        toll_r=pr+7; mc(toll_r,1,toll_r,4); ws.cell(toll_r,1).value="Toll Tax"; ws.cell(toll_r,1).font=fn(True)
                        grand_r=pr+8; mc(grand_r,1,grand_r,4)
                        ws.cell(grand_r,1).value="Grand Total"; ws.cell(grand_r,1).font=fn(True); ws.cell(grand_r,1).alignment=al()
                        cgt=ws.cell(grand_r,7); cgt.value=f"=SUM(G{pr+1}:G{grand_r-1})"; cgt.fill=yf(); cgt.alignment=al()

                    if wb.sheetnames:
                        fname = f"temp_exports/{group}_{prev_mname}_{prev_year}.xlsx"
                        wb.save(fname)
                        export_files[group] = {"path": fname, "name": f"{group}_{prev_mname}_{prev_year}.xlsx"}

                # Step 3: Reset all Google Sheets for new month
                for car, info in drivers.items():
                    reset_sheet_for_new_month(info["file_id"], info["sheet"], month, year)

                download_links = [
                    {"url": f"/download-export/{v['name']}", "label": v['name']}
                    for v in export_files.values()
                ]
                msg = (f"✅ Google Sheets reset for {month_name} {year}. "
                       f"Download the old month files below — they contain all driver entries.")
                cls = "success"

            except Exception as e:
                import traceback
                msg = f"Error: {e}\n{traceback.format_exc()}"

    now = datetime.now()
    return render_template("admin.html", msg=msg, cls=cls,
                           cur_month=now.month, cur_year=now.year,
                           download_links=download_links)

@app.route("/download-export/<filename>")
def download_export(filename):
    from flask import send_file
    # Security: only allow xlsx files from temp_exports
    if not filename.endswith(".xlsx") or "/" in filename or ".." in filename:
        return "Invalid file", 400
    path = os.path.join("temp_exports", filename)
    if not os.path.exists(path):
        return "File not found or expired", 404
    return send_file(path, as_attachment=True, download_name=filename)

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

@app.route("/ping")
def ping():
    return "ok"
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)